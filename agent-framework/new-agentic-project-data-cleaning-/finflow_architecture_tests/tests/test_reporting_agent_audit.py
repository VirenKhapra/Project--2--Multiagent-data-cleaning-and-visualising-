"""Focused, deterministic tests for the Reporting_Agent audit path (task 11.2).

These tests exercise :class:`ReportingAgent` directly (not through the
engine) and run purely against in-memory DataFrames + ``tmp_path``-backed
output directories. No real LLM is invoked: every test either supplies a
structured ``params["plan"]`` or relies on the default-plan fallback, both
of which short-circuit the optional Groq path.

The tests cover requirements 5.1 - 5.5 (input contract + result envelope),
8.1 - 8.7 / 8.10 (audit-sheet workbook + canonical filter_prep summary),
9.5 (no chart anywhere in the agent), and 10.5 (PDF rejected at the param
boundary).
"""
from __future__ import annotations

import inspect
import re
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from finflow_agent.agents.reporting_agent import (
    ReportingAgent,
    ReportingAgentParams,
)
from finflow_agent.tools.column_resolver import ColumnResolution


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------


def _xlsx_plan() -> dict:
    """Compiler-shaped plan dict for the xlsx audit-writer path."""
    return {"output_format": "xlsx", "sheet_name": "report", "title": "Report"}


def _csv_plan() -> dict:
    return {"output_format": "csv", "title": "Report"}


def _sample_df() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "id": [1, 2, 3, 4],
            "name": ["alice", "bob", "carol", "dan"],
            "amount": [10.5, 20.0, 30.25, 40.0],
        }
    )


def _filter_prep_audit_entry(op_name: str = "trim_whitespace") -> dict:
    """A non-destructive filter_prep audit-log entry shaped like the ones
    the cleaning agent emits in filter_prep mode (task 5.1).
    """
    return {
        "operation_id": f"op_{op_name}",
        "operation_type": op_name,
        "type": op_name,
        "input_row_count": 4,
        "output_row_count": 4,
        "input_columns": ["id", "name", "amount"],
        "output_columns": ["id", "name", "amount"],
        "columns_modified": ["name"],
        "warnings": [],
        "started_at": 0,
        "finished_at": 1,
        "duration_ms": 1,
        "origin": "filter_prep",
    }


def _clean_audit_entry(op_name: str = "drop_duplicates") -> dict:
    """A clean-mode audit-log entry; same shape but without the
    ``origin: filter_prep`` marker.
    """
    return {
        "operation_id": f"op_{op_name}",
        "operation_type": op_name,
        "type": op_name,
        "input_row_count": 4,
        "output_row_count": 4,
        "input_columns": ["id", "name", "amount"],
        "output_columns": ["id", "name", "amount"],
        "columns_modified": [],
        "warnings": [],
        "started_at": 0,
        "finished_at": 1,
        "duration_ms": 1,
    }


@pytest.fixture(autouse=True)
def _no_groq_key(monkeypatch):
    """Ensure the optional LLM branch never fires during these tests."""
    monkeypatch.delenv("GROQ_API_KEY", raising=False)
    yield


# ---------------------------------------------------------------------------
# 1. xlsx output produces all five canonical audit sheets
# ---------------------------------------------------------------------------


def test_reporting_agent_writes_audit_sheets_for_xlsx_output(tmp_path):
    df = _sample_df()
    audit_log = [_filter_prep_audit_entry()]
    warnings = ["sample warning from upstream"]
    column_mapping = [
        ColumnResolution(
            requested_field="amount",
            matched_column="amount",
            semantic_type="currency",
            confidence=1.0,
            reason="exact name match (case-insensitive)",
        ).model_dump()
    ]

    result = ReportingAgent().execute(
        {
            "plan": _xlsx_plan(),
            "output_dir": str(tmp_path),
            "file_prefix": "audit_workbook",
            "audit_log": audit_log,
            "warnings": warnings,
        },
        {
            "input_dataframe": df,
            "column_mapping": column_mapping,
        },
    )

    assert result.status == "success", result.error_message

    output_path = Path(result.artifacts["output_file_path"])
    assert output_path.exists()
    assert output_path.suffix == ".xlsx"
    assert output_path.parent == tmp_path

    workbook = openpyxl.load_workbook(output_path, read_only=True)
    try:
        assert workbook.sheetnames == [
            "cleaned_data",
            "audit_log",
            "warnings",
            "column_mapping",
        ]
    finally:
        workbook.close()

    assert result.artifacts["sheets_written"] == [
        "cleaned_data",
        "audit_log",
        "warnings",
        "column_mapping",
    ]


def test_reporting_agent_includes_filtered_data_sheet_when_supplied(tmp_path):
    df = _sample_df()
    filtered = df[df["amount"] > 15.0].reset_index(drop=True)

    result = ReportingAgent().execute(
        {
            "plan": _xlsx_plan(),
            "output_dir": str(tmp_path),
            "file_prefix": "audit_with_filtered",
        },
        {
            "input_dataframe": df,
            "cleaned_dataframe": df,
            "filtered_dataframe": filtered,
        },
    )

    assert result.status == "success", result.error_message
    workbook = openpyxl.load_workbook(
        Path(result.artifacts["output_file_path"]), read_only=True
    )
    try:
        assert workbook.sheetnames == [
            "cleaned_data",
            "filtered_data",
            "audit_log",
            "warnings",
            "column_mapping",
        ]
    finally:
        workbook.close()


# ---------------------------------------------------------------------------
# 2. canonical summary detection
# ---------------------------------------------------------------------------


def test_reporting_agent_canonical_summary_for_filter_prep_run(tmp_path):
    result = ReportingAgent().execute(
        {
            "plan": _xlsx_plan(),
            "output_dir": str(tmp_path),
            "file_prefix": "summary_filter_prep",
            "audit_log": [
                _filter_prep_audit_entry("trim_whitespace"),
                _filter_prep_audit_entry("normalize_column_names"),
            ],
        },
        {"input_dataframe": _sample_df()},
    )

    assert result.status == "success", result.error_message
    assert result.summary == "Data was normalized for filtering."


def test_reporting_agent_does_not_claim_cleaning_when_filter_prep_only(
    tmp_path,
):
    result = ReportingAgent().execute(
        {
            "plan": _xlsx_plan(),
            "output_dir": str(tmp_path),
            "file_prefix": "summary_no_cleaning_claim",
            "audit_log": [_filter_prep_audit_entry()],
        },
        {"input_dataframe": _sample_df()},
    )

    assert result.status == "success", result.error_message
    summary_lower = (result.summary or "").lower()
    assert "cleaning" not in summary_lower, result.summary
    assert "cleaned" not in summary_lower, result.summary


def test_reporting_agent_uses_normal_summary_for_clean_mode_run(tmp_path):
    result = ReportingAgent().execute(
        {
            "plan": _xlsx_plan(),
            "output_dir": str(tmp_path),
            "file_prefix": "summary_clean_mode",
            "audit_log": [
                _clean_audit_entry("trim_whitespace"),
                _clean_audit_entry("drop_duplicates"),
            ],
        },
        {"input_dataframe": _sample_df()},
    )

    assert result.status == "success", result.error_message
    assert result.summary != "Data was normalized for filtering."
    assert result.summary is not None
    assert result.summary.strip() != ""


def test_reporting_agent_mixed_audit_log_does_not_use_canonical_summary(
    tmp_path,
):
    """When BOTH a clean-mode entry AND a filter_prep entry are present,
    needs_cleaning was True upstream so the canonical summary must not
    fire (req 8.10: "WHERE a filter_prep step ran AND needs_cleaning was
    false ...").
    """
    result = ReportingAgent().execute(
        {
            "plan": _xlsx_plan(),
            "output_dir": str(tmp_path),
            "file_prefix": "summary_mixed",
            "audit_log": [
                _clean_audit_entry("drop_duplicates"),
                _filter_prep_audit_entry("trim_whitespace"),
            ],
        },
        {"input_dataframe": _sample_df()},
    )

    assert result.status == "success", result.error_message
    assert result.summary != "Data was normalized for filtering."


# ---------------------------------------------------------------------------
# 3. precondition: missing input_dataframe
# ---------------------------------------------------------------------------


def test_reporting_agent_returns_failed_on_missing_input_dataframe(tmp_path):
    result = ReportingAgent().execute(
        {
            "plan": _xlsx_plan(),
            "output_dir": str(tmp_path),
            "file_prefix": "missing_input",
        },
        {},
    )

    assert result.status == "failed"
    assert result.error_message == (
        "input_dataframe is required. No input dataframe provided."
    )


# ---------------------------------------------------------------------------
# 4. purity: agent must not mutate the dataframe
# ---------------------------------------------------------------------------


def test_reporting_agent_does_not_modify_dataframe(tmp_path):
    df = _sample_df()
    before_values = df.values.copy()
    before_columns = list(df.columns)
    before_index = list(df.index)

    result = ReportingAgent().execute(
        {
            "plan": _xlsx_plan(),
            "output_dir": str(tmp_path),
            "file_prefix": "no_mutation",
            "audit_log": [_filter_prep_audit_entry()],
        },
        {"input_dataframe": df},
    )

    assert result.status == "success", result.error_message

    # Caller's dataframe is unchanged in shape, dtype, and values.
    assert list(df.columns) == before_columns
    assert list(df.index) == before_index
    assert df.values.tolist() == before_values.tolist()


# ---------------------------------------------------------------------------
# 5. csv back-compat path still works
# ---------------------------------------------------------------------------


def test_reporting_agent_csv_path_still_works(tmp_path):
    df = _sample_df()

    result = ReportingAgent().execute(
        {
            "plan": _csv_plan(),
            "output_dir": str(tmp_path),
            "file_prefix": "csv_output",
        },
        {"input_dataframe": df},
    )

    assert result.status == "success", result.error_message
    output_path = Path(result.artifacts["output_file_path"])
    assert output_path.exists()
    assert output_path.suffix == ".csv"
    # The csv file contains the dataframe contents verbatim.
    written = pd.read_csv(output_path)
    pd.testing.assert_frame_equal(written, df)


def test_reporting_agent_csv_path_canonical_summary_still_applies(tmp_path):
    """The canonical summary is format-agnostic — when filter_prep ran upstream
    and the user requested a non-xlsx report, the summary still flips.
    """
    result = ReportingAgent().execute(
        {
            "plan": _csv_plan(),
            "output_dir": str(tmp_path),
            "file_prefix": "csv_canonical_summary",
            "audit_log": [_filter_prep_audit_entry()],
        },
        {"input_dataframe": _sample_df()},
    )

    assert result.status == "success", result.error_message
    assert result.summary == "Data was normalized for filtering."


# ---------------------------------------------------------------------------
# 6. PDF rejected at the param model boundary
# ---------------------------------------------------------------------------


def test_reporting_agent_pdf_rejected_at_param_validation():
    with pytest.raises(Exception):
        ReportingAgentParams.model_validate(
            {
                "plan": {"output_format": "pdf"},
            }
        )


# ---------------------------------------------------------------------------
# 7. static check: agent never imports a chart-rendering surface
# ---------------------------------------------------------------------------


_FORBIDDEN_TOKENS = (
    "add_chart",
    "insert_chart",
    "savefig",
    "pyplot",
)
_FORBIDDEN_IMPORT_PATTERNS = (
    re.compile(r"^\s*import\s+matplotlib\b", re.MULTILINE),
    re.compile(r"^\s*from\s+matplotlib\b", re.MULTILINE),
    re.compile(r"^\s*import\s+plotly\b", re.MULTILINE),
    re.compile(r"^\s*from\s+plotly\b", re.MULTILINE),
    re.compile(r"^\s*import\s+seaborn\b", re.MULTILINE),
    re.compile(r"^\s*from\s+seaborn\b", re.MULTILINE),
    re.compile(r"^\s*import\s+xlsxwriter\b", re.MULTILINE),
    re.compile(r"^\s*from\s+xlsxwriter\b", re.MULTILINE),
)


def test_reporting_agent_does_not_render_chart():
    from finflow_agent.agents import reporting_agent

    source = inspect.getsource(reporting_agent)

    for token in _FORBIDDEN_TOKENS:
        assert token not in source, (
            f"reporting_agent.py must not reference chart token {token!r}"
        )

    for pattern in _FORBIDDEN_IMPORT_PATTERNS:
        match = pattern.search(source)
        assert match is None, (
            f"reporting_agent.py must not import chart library: "
            f"{match.group(0).strip() if match else ''!r}"
        )

    # Also assert the agent never calls the chart-aggregation key from the
    # previous implementation. The replacement aggregates `audit_log` /
    # `column_mapping` instead.
    assert "chart_configs" not in source
    assert "chart_artifacts" not in source


# ---------------------------------------------------------------------------
# 8. column_mapping artifact propagates from filter agent shape
# ---------------------------------------------------------------------------


def test_reporting_agent_accepts_column_mapping_as_dicts(tmp_path):
    """The filter agent stores column_mapping entries as
    `ColumnResolution.model_dump()` dicts (task 6.1). The reporting agent
    must accept the same shape and convert them transparently.
    """
    column_mapping = [
        {
            "requested_field": "gender",
            "matched_column": "Gender",
            "semantic_type": "categorical",
            "confidence": 0.95,
            "reason": "normalized name match",
        },
        ColumnResolution(
            requested_field="age",
            matched_column="Age",
            semantic_type="numeric",
            confidence=1.0,
            reason="exact name match (case-insensitive)",
        ),
    ]

    result = ReportingAgent().execute(
        {
            "plan": _xlsx_plan(),
            "output_dir": str(tmp_path),
            "file_prefix": "column_mapping_propagation",
        },
        {
            "input_dataframe": _sample_df(),
            "column_mapping": column_mapping,
        },
    )

    assert result.status == "success", result.error_message

    workbook = openpyxl.load_workbook(
        Path(result.artifacts["output_file_path"]), read_only=True
    )
    try:
        sheet = workbook["column_mapping"]
        rows = list(sheet.values)
    finally:
        workbook.close()

    # Header row + 2 resolution rows.
    assert len(rows) == 3
    header = rows[0]
    assert header == (
        "requested_field",
        "matched_column",
        "semantic_type",
        "confidence",
        "reason",
    )
    requested_fields = {row[0] for row in rows[1:]}
    assert requested_fields == {"gender", "age"}
