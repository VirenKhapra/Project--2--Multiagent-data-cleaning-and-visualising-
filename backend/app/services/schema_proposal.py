from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

import pandas as pd

from app.services.json_safety import make_json_safe
from app.services.rule_engine import build_validation_warnings
from app.services.rule_extractor import extract_prompt_constraints, infer_semantic_constraints, merge_constraints
from app.services.semantic_schema import canonical_target_for_column
from app.services.rule_types import SEMANTIC_HINTS


COLUMN_ALIASES = {
    "date": "voucher_date",
    "entry no": "entry_no",
    "entry_number": "entry_no",
    "sub account": "sub_account",
    "ledger_name": "sub_account",
    "particulars": "details",
    "account class": "class",
    "account_class": "class",
    "account subclass": "sub_class",
    "account_subclass": "sub_class",
    "debit": "debit_amount",
    "credit": "credit_amount",
    "debit amount": "debit_amount",
    "credit amount": "credit_amount",
    "account code": "account_code",
}

def build_schema_proposal_from_file(
    path: Path,
    *,
    max_preview_rows: int,
    instruction: str = "",
) -> tuple[dict[str, Any], list[dict[str, Any]]] | None:
    extension = path.suffix.lower()
    if extension in {".csv", ".tsv", ".xlsx", ".xls"}:
        frame, total_rows = _load_tabular_preview(path, extension, max_preview_rows=max_preview_rows)
    elif extension == ".json":
        frame, total_rows = _load_json_frame(path, max_preview_rows=max_preview_rows)
    elif extension == ".txt":
        frame, total_rows = _load_text_frame(path, max_preview_rows=max_preview_rows)
    else:
        return None

    if frame is None or frame.empty:
        return None

    frame = frame.astype(object).where(pd.notnull(frame), None)
    detected_types = _infer_detected_types(frame)
    records = _records_from_frame(frame)
    source_columns = [str(column) for column in frame.columns]
    proposed_fields = [_build_field_mapping(column, detected_types.get(str(column), "string")) for column in source_columns]
    prompt_constraints = extract_prompt_constraints(source_columns, records, instruction)
    inferred_constraints = infer_semantic_constraints(source_columns, records, instruction)
    effective_constraints = merge_constraints(prompt_constraints, inferred_constraints)
    validation_warnings = build_validation_warnings(frame, effective_constraints)

    proposal = {
        "status": "awaiting_schema_approval",
        "requires_user_approval": True,
        "schema_kind": "tabular",
        "source_columns": source_columns,
        "proposed_fields": proposed_fields,
        "detected_types": detected_types,
        "preview_rows": records,
        "preview_row_count": len(records),
        "total_rows": total_rows,
        "action_schema": {
            "actions": [],
            "required_capabilities": [],
            "optional_hints": {"source": "deferred_to_agent_parser"},
            "source": "deferred_to_agent_parser",
        },
        "suggested_constraints": effective_constraints,
        "prompt_constraints": prompt_constraints,
        "validation_warnings": validation_warnings,
        "suggestion": _build_suggestion(validation_warnings),
    }
    safe_proposal = make_json_safe(proposal)
    return safe_proposal, safe_proposal.get("preview_rows", [])


def _load_tabular_preview(path: Path, extension: str, *, max_preview_rows: int) -> tuple[pd.DataFrame | None, int]:
    total_rows = 0
    if extension == ".csv":
        frame = pd.read_csv(path, nrows=max_preview_rows)
        total_rows = _count_delimited_rows(path)
    elif extension == ".tsv":
        frame = pd.read_csv(path, sep="\t", nrows=max_preview_rows)
        total_rows = _count_delimited_rows(path)
    else:
        frame = pd.read_excel(path, nrows=max_preview_rows)
        total_rows = _estimate_excel_rows(path, extension)
        if total_rows <= 0:
            total_rows = len(frame)

    frame = frame.dropna(how="all")
    frame.columns = [_normalize_source_column(column) for column in frame.columns]
    if total_rows <= 0:
        total_rows = len(frame)
    return frame, total_rows


def _load_json_frame(path: Path, *, max_preview_rows: int) -> tuple[pd.DataFrame | None, int]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    rows: list[dict[str, Any]] | None = None
    if isinstance(payload, list) and payload and all(isinstance(item, dict) for item in payload):
        rows = payload
    elif isinstance(payload, dict):
        if payload and all(not isinstance(value, (list, dict)) for value in payload.values()):
            rows = [payload]
        else:
            nested_rows = next(
                (value for value in payload.values() if isinstance(value, list) and value and all(isinstance(item, dict) for item in value)),
                None,
            )
            if nested_rows is not None:
                rows = nested_rows
    if not rows:
        return None, 0
    total_rows = len(rows)
    return pd.DataFrame(rows[:max_preview_rows]), total_rows


def _load_text_frame(path: Path, *, max_preview_rows: int) -> tuple[pd.DataFrame | None, int]:
    lines = [line.strip() for line in path.read_text(encoding="utf-8", errors="ignore").splitlines() if line.strip()]
    if not lines:
        return None, 0
    total_rows = len(lines)
    preview_lines = lines[:max_preview_rows]
    return pd.DataFrame(
        [{"line_number": index + 1, "raw_text": line} for index, line in enumerate(preview_lines)]
    ), total_rows


def _normalize_source_column(value: Any) -> str:
    raw = str(value).strip()
    alias_key = raw.lower()
    if alias_key in COLUMN_ALIASES:
        return COLUMN_ALIASES[alias_key]
    normalized = re.sub(r"[^a-z0-9]+", "_", alias_key).strip("_")
    return normalized or "column"


def _infer_detected_types(frame: pd.DataFrame) -> dict[str, str]:
    detected: dict[str, str] = {}
    for column in frame.columns:
        series = frame[column].dropna()
        if series.empty:
            detected[str(column)] = "empty"
            continue
        if pd.api.types.is_numeric_dtype(series):
            detected[str(column)] = "number"
            continue
        numeric_ratio = pd.to_numeric(series, errors="coerce").notna().mean()
        if numeric_ratio >= 0.9:
            detected[str(column)] = "number"
            continue
        if _looks_like_date_column(str(column), series):
            detected[str(column)] = "date"
        else:
            detected[str(column)] = "string"
    return detected


def _build_field_mapping(source_column: str, detected_type: str) -> dict[str, str]:
    alias_match = source_column in COLUMN_ALIASES.values()
    if alias_match:
        target = source_column
        confidence = "high"
        reason = "Matched a known finance schema alias."
    elif canonical_target_for_column(source_column):
        target = canonical_target_for_column(source_column) or source_column
        confidence = "high"
        reason = "Matched a semantic column role and mapped to the canonical target."
    elif source_column == "raw_text":
        target = "raw_text"
        confidence = "medium"
        reason = "Preserved unstructured text for user confirmation."
    else:
        target = source_column
        confidence = "medium"
        reason = "Normalized from the source header and proposed as-is."
    return {
        "source": source_column,
        "target": target,
        "detected_type": detected_type,
        "confidence": confidence,
        "reason": reason,
    }


def _records_from_frame(frame: pd.DataFrame) -> list[dict[str, Any]]:
    return frame.to_dict(orient="records")


def _count_delimited_rows(path: Path) -> int:
    try:
        with path.open("r", encoding="utf-8", errors="ignore") as handle:
            line_count = sum(1 for line in handle if line.strip())
        return max(0, line_count - 1)
    except Exception:
        return 0


def _estimate_excel_rows(path: Path, extension: str) -> int:
    try:
        if extension == ".xlsx":
            from openpyxl import load_workbook

            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet = workbook.worksheets[0] if workbook.worksheets else None
            return max(0, (sheet.max_row if sheet else 0) - 1)
        frame = pd.read_excel(path)
        return len(frame)
    except Exception:
        return 0


def _looks_like_date_column(column_name: str, series: pd.Series) -> bool:
    normalized_name = _tokenize_name(column_name)
    if normalized_name in SEMANTIC_HINTS["date_like"] or "date" in normalized_name:
        return True
    sample = series.astype(str).head(10)
    date_pattern_hits = sample.str.contains(r"\d{1,4}[-/]\d{1,2}[-/]\d{1,4}", regex=True).mean()
    return date_pattern_hits >= 0.6




def _build_suggestion(validation_warnings: list[dict[str, Any]]) -> str:
    if validation_warnings:
        return "Review the proposed schema and the flagged validation warnings before agent processing begins."
    return "Review the detected schema mapping and approve it before agent processing begins."


def _tokenize_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.strip().lower()).strip("_")
